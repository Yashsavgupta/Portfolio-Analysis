"""Service for parsing Zerodha Holdings Excel export files"""
import openpyxl
from typing import Dict, List, Optional, Tuple
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


class ZerodhaExcelParser:
    """Parse Zerodha Holdings Excel exports"""
    
    SHEET_NAMES = ['Equity', 'Combined']
    SUMMARY_FIELDS = {
        'Invested Value': 'invested_value',
        'Present Value': 'present_value',
        'Unrealized P&L': 'unrealized_pnl',
        'Unrealized P&L Pct.': 'unrealized_pnl_pct',
    }
    
    HOLDING_COLUMNS = {
        'Symbol': 'symbol',
        'ISIN': 'isin',
        'Sector': 'sector',
        'Quantity Available': 'quantity_available',
        'Quantity Discrepant': 'quantity_discrepant',
        'Quantity Long Term': 'quantity_long_term',
        'Quantity Pledged (Margin)': 'quantity_pledged_margin',
        'Quantity Pledged (Loan)': 'quantity_pledged_loan',
        'Average Price': 'average_price',
        'Previous Closing Price': 'previous_closing_price',
        'Unrealized P&L': 'unrealized_pnl',
        'Unrealized P&L Pct.': 'unrealized_pnl_pct',
    }
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.summary = {}
        self.holdings = []
        
    def parse(self) -> Tuple[Dict, List[Dict]]:
        """Parse the Excel file and return summary and holdings"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            
            # Find the right sheet
            sheet = self._get_worksheet()
            if not sheet:
                raise ValueError(f"No suitable sheet found. Expected {self.SHEET_NAMES}")
            
            # Parse summary and holdings
            self._parse_summary(sheet)
            self._parse_holdings(sheet)
            
            return self.summary, self.holdings
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")
            raise
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _get_worksheet(self):
        """Get the appropriate worksheet"""
        for sheet_name in self.SHEET_NAMES:
            if sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                # Check if sheet has data
                if sheet.max_row > 0:
                    return sheet
        return None
    
    def _parse_summary(self, sheet):
        """Parse the summary block at the top of the sheet"""
        self.summary = {
            'invested_value': 0,
            'present_value': 0,
            'unrealized_pnl': 0,
            'unrealized_pnl_pct': 0,
        }
        
        # Look for summary fields in first 20 rows
        for row in range(1, min(21, sheet.max_row + 1)):
            col_a = sheet.cell(row, 1).value
            col_b = sheet.cell(row, 2).value
            
            if col_a in self.SUMMARY_FIELDS:
                field_name = self.SUMMARY_FIELDS[col_a]
                try:
                    if field_name.endswith('_pct'):
                        value = self._parse_number(col_b)
                    else:
                        value = self._parse_currency(col_b)
                    self.summary[field_name] = value
                except (ValueError, TypeError):
                    pass
    
    def _parse_holdings(self, sheet):
        """Parse the holdings table"""
        self.holdings = []
        
        # Find the header row (row containing "Symbol")
        header_row = self._find_header_row(sheet)
        if not header_row:
            logger.warning("Could not find holdings header row")
            return
        
        # Build column index map
        header_map = self._build_header_map(sheet, header_row)
        
        # Parse data rows
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            symbol = sheet.cell(row_idx, header_map.get('Symbol', 1)).value
            
            if not symbol or str(symbol).strip() == '':
                continue
            
            holding = self._parse_holding_row(sheet, row_idx, header_map)
            if holding:
                self.holdings.append(holding)
        
        logger.info(f"Parsed {len(self.holdings)} holdings")
    
    def _find_header_row(self, sheet) -> Optional[int]:
        """Find the row containing the holdings table header"""
        for row in range(1, min(50, sheet.max_row + 1)):
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row, col).value
                if cell_value == 'Symbol':
                    return row
        return None
    
    def _build_header_map(self, sheet, header_row: int) -> Dict[str, int]:
        """Build a map of column names to column indices"""
        header_map = {}
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(header_row, col_idx).value
            if cell_value:
                header_map[cell_value] = col_idx
        return header_map
    
    def _parse_holding_row(self, sheet, row_idx: int, header_map: Dict[str, int]) -> Optional[Dict]:
        """Parse a single holding row"""
        try:
            holding = {}
            
            for excel_col, python_col in self.HOLDING_COLUMNS.items():
                col_idx = header_map.get(excel_col)
                if not col_idx:
                    continue
                
                value = sheet.cell(row_idx, col_idx).value
                
                if value is None:
                    value = 0 if 'quantity' in python_col or 'price' in python_col else None
                    holding[python_col] = value
                    continue
                
                # Parse based on column type
                if 'quantity' in python_col or 'price' in python_col or 'pnl' in python_col:
                    holding[python_col] = self._parse_number(value)
                elif 'pct' in python_col:
                    # Percentage - parse and convert to decimal
                    val = self._parse_number(value)
                    holding[python_col] = val / 100 if val else 0
                else:
                    holding[python_col] = str(value).strip() if value else None
            
            # Ensure required fields exist
            if not holding.get('symbol'):
                return None
            
            # Calculate derived fields
            holding['quantity'] = holding.get('quantity_available', 0) or 0
            holding['quantity_available'] = holding.get('quantity_available', holding['quantity']) or holding['quantity']
            holding['quantity_long_term'] = holding.get('quantity_long_term', 0) or 0
            holding['quantity_pledged_margin'] = holding.get('quantity_pledged_margin', 0) or 0
            holding['quantity_pledged_loan'] = holding.get('quantity_pledged_loan', 0) or 0
            holding['average_price'] = holding.get('average_price', 0) or 0
            holding['previous_closing_price'] = holding.get('previous_closing_price', 0) or 0
            holding['invested_value'] = round(holding['quantity'] * holding['average_price'], 2)
            holding['market_value'] = round(holding['quantity'] * holding['previous_closing_price'], 2)

            if holding.get('unrealized_pnl') is None:
                holding['unrealized_pnl'] = round(holding['market_value'] - holding['invested_value'], 2)

            if holding.get('unrealized_pnl_pct') in (None, 0):
                invested_value = holding['invested_value']
                holding['unrealized_pnl_pct'] = round(
                    (holding['unrealized_pnl'] / invested_value * 100) if invested_value > 0 else 0,
                    2,
                )
            
            return holding
            
        except Exception as e:
            logger.warning(f"Error parsing holding at row {row_idx}: {str(e)}")
            return None
    
    def _parse_currency(self, value) -> float:
        """Parse a currency value, handling ₹ symbol and commas"""
        if value is None:
            return 0.0
        
        if isinstance(value, (int, float)):
            return float(value)
        
        # Handle string values
        value_str = str(value).strip()
        value_str = value_str.replace('₹', '').replace(',', '').strip()
        
        return float(value_str)
    
    def _parse_number(self, value) -> float:
        """Parse a numeric value"""
        if value is None:
            return 0.0
        
        if isinstance(value, (int, float)):
            return float(value)
        
        # Handle string values
        value_str = str(value).strip()
        value_str = value_str.replace('%', '').replace(',', '').strip()
        
        return float(value_str) if value_str else 0.0
